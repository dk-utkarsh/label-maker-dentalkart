import os
import io
import zipfile
import tempfile
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from reportlab.lib.units import mm, inch
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader, simpleSplit
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.barcode import code128
from openpyxl import load_workbook
from PIL import Image
import barcode
from barcode.writer import ImageWriter

app = FastAPI(title="Label Generator")

# Use /tmp on Vercel (read-only filesystem), local uploads/ otherwise
if os.environ.get("VERCEL"):
    UPLOAD_DIR = Path("/tmp/uploads")
else:
    UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# ──────────────────────────────────────────────
# Routes
# ──────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index():
    html_path = Path(__file__).parent / "static" / "index.html"
    return html_path.read_text()


@app.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    """Read Excel, return column names and first 5 rows as preview."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an Excel file (.xlsx or .xls)")

    content = await file.read()
    filepath = UPLOAD_DIR / "data.xlsx"
    filepath.write_bytes(content)

    columns, rows = read_excel(filepath)
    if not columns:
        raise HTTPException(400, "No data found in the Excel file")

    preview_rows = rows[:5]
    return JSONResponse({
        "columns": columns,
        "total_rows": len(rows),
        "preview": preview_rows,
    })


@app.post("/upload-logo")
async def upload_logo(file: UploadFile = File(...)):
    """Save uploaded logo and return its path."""
    ext = Path(file.filename).suffix.lower()
    if ext not in (".png", ".jpg", ".jpeg", ".svg", ".webp", ".bmp"):
        raise HTTPException(400, "Unsupported image format")

    content = await file.read()
    logo_path = UPLOAD_DIR / f"logo{ext}"
    # Remove old logos
    for old in UPLOAD_DIR.glob("logo.*"):
        old.unlink()
    logo_path.write_bytes(content)
    return {"status": "ok", "path": str(logo_path)}


@app.post("/generate")
async def generate_labels(
    width: float = Form(...),
    height: float = Form(...),
    mode: str = Form("single"),  # "single" or "zip"
    field_config: str = Form(...),  # JSON string
    filename_field: str = Form(""),  # column to use for ZIP filenames
    barcode_field: str = Form(""),  # column to generate barcode from
    banner_text: str = Form(""),  # text for bottom banner
):
    """Generate PDF labels and return as download."""
    import json

    excel_path = UPLOAD_DIR / "data.xlsx"
    if not excel_path.exists():
        raise HTTPException(400, "No Excel file uploaded yet")

    # Find logo
    logo_path = None
    for ext in (".png", ".jpg", ".jpeg", ".webp", ".bmp"):
        p = UPLOAD_DIR / f"logo{ext}"
        if p.exists():
            logo_path = p
            break

    columns, rows = read_excel(excel_path)
    if not rows:
        raise HTTPException(400, "No data rows found in Excel")

    try:
        config = json.loads(field_config)
    except json.JSONDecodeError:
        raise HTTPException(400, "Invalid field configuration")

    # config is a list of {column, role} where role is title/body/footer/hidden
    active_fields = [f for f in config if f.get("role") != "hidden"]

    if mode == "zip":
        return generate_zip(rows, active_fields, width, height, logo_path, filename_field, barcode_field, banner_text)
    else:
        return generate_single_pdf(rows, active_fields, width, height, logo_path, barcode_field, banner_text)


# ──────────────────────────────────────────────
# Excel Reading
# ──────────────────────────────────────────────

def read_excel(filepath: Path):
    """Read Excel file, return (columns, rows) where rows is list of dicts."""
    wb = load_workbook(str(filepath), data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for cell in ws[1]:
            val = str(cell.value).strip() if cell.value is not None else ""
            headers.append(val)

        if not any(headers):
            continue

        # Clean empty trailing headers
        while headers and not headers[-1]:
            headers.pop()

        if not headers:
            continue

        rows = []
        for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            row_data = {}
            all_empty = True
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    cell_val = format_cell(val)
                    row_data[headers[i]] = cell_val
                    if cell_val:
                        all_empty = False
            if not all_empty:
                rows.append(row_data)

        if rows:
            return headers, rows

    return [], []


def format_cell(val):
    """Format cell value to string."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:.2f}"
    return str(val).strip()


# ──────────────────────────────────────────────
# PDF Generation
# ──────────────────────────────────────────────

def generate_single_pdf(rows, fields, width_mm, height_mm, logo_path, barcode_field="", banner_text=""):
    """Generate a single PDF with all labels (one per page)."""
    buf = io.BytesIO()
    w = width_mm * mm
    h = height_mm * mm

    c = canvas.Canvas(buf, pagesize=(w, h))

    for i, row in enumerate(rows):
        draw_label(c, row, fields, w, h, width_mm, height_mm, logo_path, barcode_field, banner_text)
        if i < len(rows) - 1:
            c.showPage()

    c.save()
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=labels.pdf"}
    )


def generate_zip(rows, fields, width_mm, height_mm, logo_path, filename_field, barcode_field="", banner_text=""):
    """Generate individual PDFs in a ZIP."""
    zip_buf = io.BytesIO()

    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, row in enumerate(rows):
            pdf_buf = io.BytesIO()
            w = width_mm * mm
            h = height_mm * mm
            c = canvas.Canvas(pdf_buf, pagesize=(w, h))
            draw_label(c, row, fields, w, h, width_mm, height_mm, logo_path, barcode_field, banner_text)
            c.save()
            pdf_buf.seek(0)

            # Filename from field or fallback
            name = ""
            if filename_field and filename_field in row:
                name = row[filename_field]
            if not name:
                name = f"label_{i + 1}"
            # Sanitize filename
            safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(name))
            zf.writestr(f"{safe_name}.pdf", pdf_buf.read())

    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=labels.zip"}
    )



def generate_barcode_image(value):
    """Generate a barcode image from a value and return as PIL Image."""
    try:
        code128 = barcode.get_barcode_class('code128')
        bc = code128(str(value), writer=ImageWriter())
        buf = io.BytesIO()
        bc.write(buf, options={"write_text": False, "module_height": 8, "quiet_zone": 2})
        buf.seek(0)
        return Image.open(buf)
    except Exception:
        return None


def draw_label(c, data, field_config, width, height, w_mm, h_mm, logo_path, barcode_field="", banner_text=""):
    """Draw a single label on the canvas with high precision matching the web preview."""
    # Match index.html padding (Math.max(8, wPx * 0.045))
    # Note: frontend width in px is w_mm * scale. We use points (mm * 2.83)
    pad = max(8, width * 0.045)
    
    cursor = [height - pad]
    x_offset = pad
    content_width = width - (2 * pad)
    
    # Font scale factor to match browser rendering
    FONT_SCALE = 0.82

    def get_font_pt(role, size_str="auto"):
        base = 10
        if role == "title": base = 13
        if role == "footer": base = 8
        scales = {"xs": 0.7, "sm": 0.85, "md": 1.0, "lg": 1.25, "xl": 1.6, "auto": 1.0}
        return base * scales.get(size_str, 1.0) * FONT_SCALE

    # Logo
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image.open(logo_path)
            img_w, img_h = img.size
            max_w = content_width * 0.30
            max_h = height * 0.14
            scale = min(max_w/img_w, max_h/img_h)
            draw_w, draw_h = img_w * scale, img_h * scale
            c.drawImage(logo_path, x_offset, cursor[0] - draw_h, width=draw_w, height=draw_h, mask='auto', preserveAspectRatio=True)
            cursor[0] -= (draw_h + pad * 0.5)
        except: pass
    else:
        cursor[0] -= (pad * 0.4)

    active = [f for f in field_config if f.get("role") != "hidden"]
    titles = [f for f in active if f.get("role") == "title"]
    bodies = [f for f in active if f.get("role") == "body"]
    footers = [f for f in active if f.get("role") == "footer"]

    def draw_field_group(fields, role):
        if not fields: return
        
        rows = []
        curr_row = []
        for f in fields:
            if f.get("sameRow") and curr_row:
                curr_row.append(f)
            else:
                if curr_row: rows.append(curr_row)
                curr_row = [f]
        if curr_row: rows.append(curr_row)
        
        for row_fields in rows:
            # Filter fields with content or border
            active_row_fields = [f for f in row_fields if data.get(f['column']) or f.get('border')]
            if not active_row_fields: continue
            
            # 1. First, compute the required height of the row by wrapping each field
            cell_w = content_width / len(active_row_fields)
            row_data = [] # List of {lines: [], fs: pt, align: str, border: bool, ...}
            
            max_row_h = 0
            any_row_border = any(f.get("border") for f in active_row_fields)

            for f in active_row_fields:
                val = str(data.get(f['column'], ""))
                fs = get_font_pt(role, f.get("fontSize", "auto"))
                if val: 
                    val = f.get("prefix", "") + val + f.get("suffix", "")
                    if f.get("uppercase"): val = val.upper()
                    if f.get("showLabel"): val = f['column'] + ": " + val
                
                # Check for border padding/margins
                target_w = cell_w - (8 if f.get("border") or any_row_border else 0)
                
                # Font metrics for wrapping
                font_name = "Helvetica-Bold" if role=="title" or f.get("bold") else "Helvetica"
                lines = simpleSplit(val, font_name, fs, target_w) if val else []
                if not lines and (f.get("border") or any_row_border): lines = [""] # for placeholder borders
                
                lh = fs * 1.35
                cell_h = len(lines) * lh + (4 if f.get("border") or any_row_border else 0)
                max_row_h = max(max_row_h, cell_h)
                
                row_data.append({
                    "lines": lines,
                    "fs": fs,
                    "font_name": font_name,
                    "align": f.get("align", "left"),
                    "border": f.get("border") or any_row_border,
                    "cell_w": cell_w,
                    "target_w": target_w,
                    "line_h": lh
                })
            
            if cursor[0] - max_row_h < pad: break
            
            # 2. Draw the row fields
            curr_x = x_offset
            for f_data in row_data:
                # Border
                if f_data["border"]:
                    c.setStrokeColorRGB(0.8, 0.8, 0.8)
                    c.setLineWidth(0.5)
                    c.rect(curr_x, cursor[0] - max_row_h, f_data["cell_w"], max_row_h, stroke=1, fill=0)
                
                # Text
                c.setFont(f_data["font_name"], f_data["fs"])
                c.setFillColorRGB(0,0,0)
                
                # Center vertically? For now let's just draw from top like HTML
                txt_y = cursor[0] - f_data["fs"] - (4 if f_data["border"] else 0)
                draw_x_base = curr_x + (4 if f_data["border"] else 0)
                
                for line in f_data["lines"]:
                    if f_data["align"] == "center":
                        c.drawCentredString(draw_x_base + f_data["target_w"]/2, txt_y, line)
                    elif f_data["align"] == "right":
                        c.drawRightString(draw_x_base + f_data["target_w"], txt_y, line)
                    else:
                        c.drawString(draw_x_base, txt_y, line)
                    txt_y -= f_data["line_h"]
                
                curr_x += f_data["cell_w"]
            
            cursor[0] -= (max_row_h + (0 if any_row_border else pad * 0.15))

    # 1. Titles
    draw_field_group(titles, "title")
    
    # 2. Barcode
    if barcode_field and data.get(barcode_field):
        try:
            bc_val = str(data.get(barcode_field))
            bc_h = height * 0.10
            bc = code128.Code128(bc_val, barHeight=bc_h, barWidth=1.5)
            bc_draw_w = bc.width
            c.saveState()
            c.translate((width - bc_draw_w)/2, cursor[0] - bc_h)
            bc.drawOn(c, 0, 0)
            c.restoreState()
            cursor[0] -= (bc_h + pad * 0.4)
        except: pass

    # 3. Bodies
    draw_field_group(bodies, "body")
    
    # 4. Footers
    if footers:
        has_content = any(data.get(f['column']) or f.get('border') for f in footers)
        if has_content:
            footer_area_h = sum([get_font_pt("footer", f.get("fontSize", "auto")) * 1.5 for f in footers]) + pad
            if cursor[0] > pad + footer_area_h:
                cursor[0] = pad + footer_area_h
            
            if not any(f.get("border") for f in footers):
                c.setStrokeColorRGB(0.9, 0.9, 0.9)
                c.line(pad, cursor[0] + 2, width-pad, cursor[0] + 2)
            draw_field_group(footers, "footer")

    # 5. Banner
    if banner_text:
        banner_fs = max(6, height * 0.025)
        c.setFillColorRGB(0.06, 0.06, 0.06)
        c.rect(0, 0, width, banner_fs * 2, stroke=0, fill=1)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", banner_fs)
        c.drawCentredString(width/2, banner_fs * 0.6, banner_text.upper())
    c.setFillColorRGB(0, 0, 0)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=5050, reload=True)
